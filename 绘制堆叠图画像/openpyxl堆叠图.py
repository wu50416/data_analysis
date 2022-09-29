from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from collections import OrderedDict

from openpyxl.chart.axis import ChartLines
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart import BarChart, Reference, label as chart_label
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as ChartFont
from openpyxl.drawing.line import LineProperties



class PictureHandler(object):
    def __init__(self, input, output, origin_sheet='源数据', picture_sheet='画像'):
        self.input = input
        self.output = output

        self.origin_sheet = origin_sheet
        self.picture_sheet = picture_sheet

    # 设置图像格式(这部分可直接忽略,主要看下面的add_chart函数)
    def create_bar_chart(self,sheet, begin_row, end_row):
        chart = BarChart()
        chart.height = 7
        chart.width = 14
        chart.type = "col"
        chart.style = 2
        chart.title = sheet.cell(row=begin_row, column=1).value
        # chart1.y_axis.title = 'Test number'
        # chart1.x_axis.title = 'Sample length (mm)'

        # 图表数据
        data = Reference(sheet, min_col=sheet.min_column + 1, max_col=sheet.max_column, min_row=begin_row,
                         max_row=end_row)
        # from_rows: 当值为True时，一行是一个系列，为False时，一列是一个系列
        # titles_from_data: 当值为True时，data包含系列名称
        chart.add_data(data, titles_from_data=True, from_rows=True)

        # x轴标签
        cats = Reference(sheet, min_col=sheet.min_column + 2, max_col=sheet.max_column, min_row=sheet.min_row,
                         max_row=sheet.min_row)
        chart.set_categories(cats)

        # 添加数字标签
        chart.dLbls = chart_label.DataLabelList()
        chart.dLbls.showVal = True

        chart.grouping = "percentStacked"  # 设置为百分比堆叠图
        chart.overlap = 100  # 当为堆叠图时需要设置为100
        chart.y_axis.scaling.min = 0  # 设置y轴最大值为0
        chart.y_axis.scaling.max = 1  # 设置y轴最大值为1

        font_test = ChartFont(typeface='微软雅黑')
        cp = CharacterProperties(latin=font_test)
        chart.y_axis.textProperties = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
        chart.x_axis.textProperties = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
        chart.legend.textProperties = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
        chart.dLbls.textProperties = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
        chart.y_axis.delete = True  # 设置隐藏y轴标签

        # 设置隐藏网格线
        chart.y_axis.majorGridlines = ChartLines()
        sgp = GraphicalProperties(ln=LineProperties(noFill=True))
        chart.y_axis.majorGridlines.spPr = sgp

        return chart

    def add_chart(self):
        workbook = load_workbook(self.output)
        sheet = workbook[self.picture_sheet]

        # 根据第一列标签名称，统计各标签分别有多少行数据
        label_dict = OrderedDict()
        for row_num in range(2, sheet.max_row + 1):
            label = sheet.cell(row=row_num, column=1).value
            label_dict[label] = label_dict.get(label, 0) + 1

        # 生成图表
        begin_row = 2
        for index, (label, row_num) in enumerate(label_dict.items()):
            # 确定标签的数据起始行、结束行
            end_row = begin_row + row_num - 1
            chart = self.create_bar_chart(sheet, begin_row=begin_row, end_row=end_row)
            begin_row += row_num

            # 确定图表放置位置
            chart_row_num = 14 * (index // 2) + 1
            if index % 2 == 0:
                chart_col_num = sheet.max_column + 2
            else:
                chart_col_num = sheet.max_column + 10
            chart_position = get_column_letter(chart_col_num) + str(chart_row_num)
            sheet.add_chart(chart, chart_position)

        workbook.save(self.output)


    def start(self):
        self.add_chart()

if __name__ == '__main__':
    input_file = f'test_input.xlsx'
    output_file = f'test_input.xlsx'
    picture_handler = PictureHandler(input=input_file, output=output_file)
    picture_handler.start()

